from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from dateutil.relativedelta import relativedelta

class ExcelSheet:
    def __init__(self,partner_name,attributes,partner_data) -> None:
        self.partner_name = partner_name
        self.attributes = attributes
        self.partner_data = partner_data
        self.workbook = load_workbook(filename='sample_report.xlsx')
        self.sheet = self.workbook.active
        
    def get_status(self,data):
        res = status = ""
        if data < 0:
            data = data / -1
            res = "Drop " +  str(data) + " %"
            status = "drop"
        else:
            res = "Rise " + str(data) + " %"
            status = 'rise'
        return res,status
        
    def update_excel_cols(self):
        row_key = 2
        for row in range(11):
            partner = self.partner_name[row]
            if partner in self.partner_data.keys():
                data = self.partner_data[partner]
                utc_date = datetime.strptime(data['utc_date'], '%Y-%m-%d')
                prev_utc_date = (utc_date - relativedelta(days=1)).strftime('%Y-%m-%d')
                prev_yr_utc_date = (utc_date - relativedelta(years=1)).strftime('%Y-%m-%d')

                self.report_date = utc_date.strftime('%d %b %Y')

                powerbi_utc_date = datetime.strptime(prev_utc_date,'%Y-%m-%d')
                powerbi_prev_yr_utc_date = (powerbi_utc_date - relativedelta(years=1)).strftime('%Y-%m-%d')
                utc_date = utc_date.strftime('%Y-%m-%d')

                col_key = 'B'

                cell_location = col_key + str(row_key)
                self.sheet[cell_location] = "Imp Count For " + utc_date
                cell_location = chr(ord(col_key) + 1) + str(row_key)
                self.sheet[cell_location] = "Imp Count For " + prev_utc_date
                cell_location = chr(ord(col_key) + 2) + str(row_key)
                self.sheet[cell_location] = "Imp Count For " + prev_yr_utc_date
                
                if row in [0,1,2,3,5,10]:
                    cell_location = chr(ord(col_key) + 5) + str(row_key)
                    self.sheet[cell_location] = "Imp Count For " + utc_date
                    cell_location = chr(ord(col_key) + 6) + str(row_key)
                    self.sheet[cell_location] = "Imp Count For " + prev_utc_date

                if row in [0,1,2,3,4,5,10]:
                    if row == 4:
                        cell_location = chr(ord(col_key) + 8) + str(row_key)
                        self.sheet[cell_location] = "Imp Count For " + utc_date
                        cell_location = chr(ord(col_key) + 9) + str(row_key)
                        self.sheet[cell_location] = "Imp Count For " + prev_yr_utc_date
                    else:
                        cell_location = chr(ord(col_key) + 8) + str(row_key)
                        self.sheet[cell_location] = "Imp Count For " + prev_utc_date
                        cell_location = chr(ord(col_key) + 9) + str(row_key)
                        self.sheet[cell_location] = "Imp Count For " + powerbi_prev_yr_utc_date

            row_key += 4

    def update_impression_count(self):
        yellow = "00FFFF00"
        row_key = 3
        for row in range(11):
            partner = self.partner_name[row]
            if partner in self.partner_data.keys():
                data = self.partner_data[partner]
                col_key = 'B'
                if row == 6 or row == 7 or row == 8 or row == 9:
                    tmp = 5
                else:
                    tmp = 11
                for cols in range(tmp):
                    field = self.attributes[cols]
                    cell_location = col_key + str(row_key)

                    # for dod and yoy drop
                    if cols in [3,4,7,10]:
                        if partner == 'Twitter' and cols == 7:
                            cell_value = str(data[field])
                        else:
                            cell_value = status = ""
                            if data[field] == "NA":
                                cell_value = "NA"
                            else:
                                cell_value,status = self.get_status(data[field])

                            self.sheet[cell_location] = cell_value
                            if status == 'drop' and data[field] < -20:
                                self.sheet[cell_location].fill = PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
                    # for impression count 
                    else:
                        cell_value = str(data[field])
                        self.sheet[cell_location] = cell_value

                    col_key = chr(ord(col_key) + 1)

            row_key += 4

    def create_excel_report(self):
        # update the columns values
        self.update_excel_cols()

        # update impression counts
        self.update_impression_count()

        self.name = "Metrics for Walled Garden Partners for " + self.report_date + ".xlsx"
        self.workbook.save(filename=self.name)
        return self.name

# partner_name = ['Pinterest','Linkedin','Spotify','Snapchat','Twitter','Facebook','Youtube - Google Ads','Youtube - DV 360','Youtube - Partner Sold','Youtube - Reserve','Yahoo']
# attributes = ['snowflake_imp','previous_day_snowflake_imp','previous_yr_snowflake_imp','snowflake_DoD_drop','snowflake_YoY_drop','athena_imp','previous_day_athena_imp','athena_DoD_drop','powerbi_imp','previous_yr_powerbi_imp','powerbi_YoY_drop']
# partner_data = {"Pinterest": 
#         {
#             "utc_date": "2022-12-01",
#             "snowflake_imp": 3154448608,
#             "previous_yr_snowflake_imp": 2775917869,
#             "snowflake_YoY_drop": 13.64,
#             "previous_day_snowflake_imp": 3913536537,
#             "snowflake_DoD_drop": -19.4,
#             "athena_imp": 0,
#             "previous_day_athena_imp": 3913536537,
#             "athena_DoD_drop": 0,
#             "powerbi_imp": 0,
#             "previous_yr_powerbi_imp": 0,
#             "powerbi_YoY_drop": 0
#         },
#         "Linkedin":
#             {
#                 "utc_date": "2022-12-01",
#                 "snowflake_imp": 3154448608,
#                 "previous_yr_snowflake_imp": 2775917869,
#                 "snowflake_YoY_drop": 13.64,
#                 "previous_day_snowflake_imp": 3913536537,
#                 "snowflake_DoD_drop": -19.4,
#                 "athena_imp": 0,
#                 "previous_day_athena_imp": 3913536537,
#                 "athena_DoD_drop": 0,
#                 "powerbi_imp": 0,
#                 "previous_yr_powerbi_imp": 0,
#                 "powerbi_YoY_drop": 0
#             },
#         "Spotify":
#             {
#                 "utc_date": "2022-12-01",
#                 "snowflake_imp": 3154448608,
#                 "previous_yr_snowflake_imp": 2775917869,
#                 "snowflake_YoY_drop": 13.64,
#                 "previous_day_snowflake_imp": 3913536537,
#                 "snowflake_DoD_drop": -19.4,
#                 "athena_imp": 0,
#                 "previous_day_athena_imp": 3913536537,
#                 "athena_DoD_drop": 0,
#                 "powerbi_imp": 0,
#                 "previous_yr_powerbi_imp": 0,
#                 "powerbi_YoY_drop": 0
#             },
#         "Snapchat":
#             {
#                 "utc_date": "2022-12-01",
#                 "snowflake_imp": 3154448608,
#                 "previous_yr_snowflake_imp": 2775917869,
#                 "snowflake_YoY_drop": 13.64,
#                 "previous_day_snowflake_imp": 3913536537,
#                 "snowflake_DoD_drop": -19.4,
#                 "athena_imp": 0,
#                 "previous_day_athena_imp": 3913536537,
#                 "athena_DoD_drop": 0,
#                 "powerbi_imp": 0,
#                 "previous_yr_powerbi_imp": 0,
#                 "powerbi_YoY_drop": 0
#             },
#         "Twitter":
#             {
#                 "utc_date": "2022-12-01",
#                 "snowflake_imp": 3154448608,
#                 "previous_yr_snowflake_imp": 2775917869,
#                 "snowflake_YoY_drop": 13.64,
#                 "previous_day_snowflake_imp": 3913536537,
#                 "snowflake_DoD_drop": -19.4,
#                 "athena_imp": '',
#                 "previous_day_athena_imp": '',
#                 "athena_DoD_drop": '',
#                 "powerbi_imp": 0,
#                 "previous_yr_powerbi_imp": 0,
#                 "powerbi_YoY_drop": 0
#             },
#         "Facebook":
#             {
#                 "utc_date": "2022-11-30",
#                 "snowflake_imp": 3154448608,
#                 "previous_yr_snowflake_imp": 2775917869,
#                 "snowflake_YoY_drop": 13.64,
#                 "previous_day_snowflake_imp": 3913536537,
#                 "snowflake_DoD_drop": -19.4,
#                 "athena_imp": 0,
#                 "previous_day_athena_imp": 3913536537,
#                 "athena_DoD_drop": 0,
#                 "powerbi_imp": 0,
#                 "previous_yr_powerbi_imp": 0,
#                 "powerbi_YoY_drop": 0
#             },
#         "Youtube - Google Ads": {
#             "utc_date": "2022-12-01",
#             "snowflake_imp": 3154448608,
#             "previous_yr_snowflake_imp": 2775917869,
#             "snowflake_YoY_drop": 13.64,
#             "previous_day_snowflake_imp": 3913536537,
#             "snowflake_DoD_drop": -19.4,
#         },
#         "Youtube - DV 360": {
#             "utc_date": "2022-12-01",
#             "snowflake_imp": 3154448608,
#             "previous_yr_snowflake_imp": 2775917869,
#             "snowflake_YoY_drop": 13.64,
#             "previous_day_snowflake_imp": 3913536537,
#             "snowflake_DoD_drop": -19.4,
#         },
#         "Youtube - Partner Sold": {
#             "utc_date": "2022-12-01",
#             "snowflake_imp": 3154448608,
#             "previous_yr_snowflake_imp": 2775917869,
#             "snowflake_YoY_drop": 13.64,
#             "previous_day_snowflake_imp": 3913536537,
#             "snowflake_DoD_drop": -19.4,
#         },
#         "Youtube - Reserve": {
#             "utc_date": "2022-12-01",
#             "snowflake_imp": 3154448608,
#             "previous_yr_snowflake_imp": 2775917869,
#             "snowflake_YoY_drop": 13.64,
#             "previous_day_snowflake_imp": 3913536537,
#             "snowflake_DoD_drop": -19.4,
#         },
#         "Yahoo":
#             {
#                 "utc_date": "2022-12-01",
#                 "snowflake_imp": 3154448608,
#                 "previous_yr_snowflake_imp": 2775917869,
#                 "snowflake_YoY_drop": 13.64,
#                 "previous_day_snowflake_imp": 3913536537,
#                 "snowflake_DoD_drop": -19.4,
#                 "athena_imp": 0,
#                 "previous_day_athena_imp": 3913536537,
#                 "athena_DoD_drop": 0,
#                 "powerbi_imp": 0,
#                 "previous_yr_powerbi_imp": 0,
#                 "powerbi_YoY_drop": 0
#             }
# }

# report = ExcelSheet(partner_name,attributes,partner_data)
# report.create_excel_report()
